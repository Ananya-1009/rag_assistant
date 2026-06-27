from pathlib import Path
from openpyxl import load_workbook
def extract_excel(file_path:Path)->str:
    workbook=load_workbook(filename=file_path,data_only=True)
    text=[]
    for sheet in workbook.worksheets:
        text.append(f"Sheet:{sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values=[
                str(cell)
                if cell is not None
                else ""
                for cell in row
            ]
            text.append(" | ".join(values))
        text.append("")
    workbook.close()
    return "\n".join(text).strip()
